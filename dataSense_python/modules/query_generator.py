import json
import pandas as pd
import openpyxl
import os

# Function to save DDL statements into the existing Excel workbook
def save_ddl_to_excel(workbook, sheet_name, ddl_dataframe):
    from config.logger import logger
    # Create a new sheet with the specified name
    new_sheet = workbook.create_sheet(title=sheet_name)
    
    # Write the DataFrame to the new sheet
    for r_idx, row in ddl_dataframe.iterrows():
        for c_idx, value in enumerate(row):
            new_sheet.cell(row=r_idx + 2, column=c_idx + 1, value=value)

    # Write the column headers
    for col_idx, col_name in enumerate(ddl_dataframe.columns):
        new_sheet.cell(row=1, column=col_idx + 1, value=col_name)

    logger.info(f"New sheet has been created, sheet name : {sheet_name}")
    

def generate_ddl_prompt(domain,table_name,db_type,db_version,column_details):
    return f"""
    Please generate a SQL script as per industry standards to create the following table {table_name} in {db_type}{db_version} without using open datatypes and increase datatype length by 20% and roundoff the length to multiples of 5,also include the necessary constraints without length check:
        {column_details}
    """
def save_sql_file(sql_file_path,response):
    try:
        from config.logger import logger
        with open(sql_file_path, 'w') as file:
            file.write(response+ '\n') 
            logger.info(f"SQL File successfully created at {sql_file_path}")
    except IOError as e:
        logger.info(f"IOError saving {sql_file_path}: {e}")
    except Exception as e:
        logger.info(f"Error saving {sql_file_path}: {e}")

     

def query_generator(domain,file_name,output_excel_path,output_path,db_name):
    try:
        from config.logger import logger
        from core.gemini_ai import gen_ai_chat
        from generals.general import extract_create_table_query,parse_sql_query,consolidate_constraints,get_file_name_without_extension,classify_constraints,create_directory
        df = pd.read_excel(output_excel_path, sheet_name='Profile')
        # Convert dataframe to string representation for the prompt
        json_data = df.to_json(orient='records', date_format='iso')
        json_data = json.loads(json_data)
        column_details = json.dumps(json_data, indent=4)
        table_name = file_name  # Replace with your table name
        # Example database types and versions
        workbook = openpyxl.load_workbook(output_excel_path)
        final_output_data = {}
        final_output_data[file_name] = []
        prompt = generate_ddl_prompt(domain,table_name,db_name,'',column_details)
        response = extract_create_table_query(gen_ai_chat(prompt))
        if response:
            sql_output_base_path = os.path.join(output_path,get_file_name_without_extension(file_name),'sql_queries')
            sql_file_path = os.path.join(create_directory(sql_output_base_path),f"{get_file_name_without_extension(file_name)}_{db_name.lower()}.sql") 
            final_output_data[file_name].append({
                'title':f'{db_name}',
                # 'sql_file_path':sql_file_path,
                'query':response,
            })

            
            save_sql_file(sql_file_path,response)
            # Parse the SQL query
            columns, constraints = parse_sql_query(response)

            # Consolidate constraints
            consolidated_data = consolidate_constraints(columns, constraints)

            # Classify and create the final DataFrame
            consolidated_df = classify_constraints(consolidated_data)
            # Add logic to store each DDL into the existing Excel
            sheet_name = f"{db_name.lower()}_ddl"
            save_ddl_to_excel(workbook, sheet_name, consolidated_df)
            
        
        # Save the workbook with the new sheets added
        workbook.save(output_excel_path)
        logger.info(f"file has been updated with DDL statement at {output_excel_path}")
        return final_output_data
    except Exception as e:
        logger.info(f"Error in process_generator : {str(e)} ")
        return False
    